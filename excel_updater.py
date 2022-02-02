from os import truncate
import openpyxl
import sqlite3
import json
import requests
from lxml import html
from openpyxl.worksheet.table import Table
from parse import search_stock_api, update_info, ratio_columns
from compare import compare

conn = sqlite3.connect('base.db')
curs = conn.cursor()

session = requests.Session()
def tryFloat(value):
    try:
        return float(value)
    except:
        return value
def insert_data(left_ltr, first_num, right_ltr, last_num, data, ws):
    for i in range(first_num,last_num+1):
        ws.merge_cells(f'{left_ltr}{i}:{right_ltr}{i}')
        ws[f'{left_ltr}{i}']=tryFloat(data[i-first_num])
def next_ltr(s,n=-1):
    if s[n]=='z' and n%len(s)==0:
        s='a'*(-n+1)
    else:
        s = list(s)
        s[n]=chr((ord(s[n])-96)%26+97)
        s = ''.join(s)
        if s[n]=='a':
            s=next_ltr(s,n-1)
    return s

def add_table(left_ltr, upper_num, table, ws):
    #print(table)
    #input()
    for column in table:
    #    print(left_ltr)
        insert_data(left_ltr, upper_num, left_ltr, upper_num+len(column)-1, column, ws)
        left_ltr = next_ltr(left_ltr)

def update_excel_file(filename, tickers):
    wb = openpyxl.load_workbook('book2.xlsx')
    ws = wb.worksheets[0]

    ws.merge_cells(f'b2:b2')
    ws['b2']=search_stock_api(tickers[0],'n')
    
    update_info(tickers[0])
    data_other_info = list(curs.execute(f'SELECT "Market Cap", "Shares Outstanding", "Employees" from [{tickers[0]}_other_info] ').fetchone())
    profile_tree = html.fromstring(session.get(f'https://stockanalysis.com/stocks/{tickers[0]}/company/').content)
    profile_text_list = profile_tree.xpath('//*[@id="main"]/div/div[2]/div[2]/div[1]/table/tbody//text()')

    founded = int(profile_text_list[profile_text_list.index('Founded')+1])
    data_other_info.insert(1,profile_tree.xpath('//*[@id="main"]/div/div[1]/section/div[1]//text()')[0])
    data_other_info.insert(3,profile_tree.xpath('//*[@id="main"]/div/div[2]/div[2]/div[5]/table/tbody/tr[2]/td[2]/text()')[0])
    data_other_info.insert(4,founded)
    
    for i in range(4,10):
        ws.merge_cells(f'b{i}:d{i}')
        ws[f'b{i}']=data_other_info[i-4]
    insert_data('b',4,'d',9,data_other_info,ws)
    
    ratio_row = curs.execute(f'SELECT * FROM [{tickers[0]}_ratio]').fetchone()

    insert_data('h',4,'j',6,ratio_columns,ws)
    insert_data('k',4,'k',6,ratio_row,ws)

    insert_data('o',4,'q',11,ratio_columns[2:],ws)
    insert_data('r',4,'r',11,ratio_row[2:],ws)

    insert_data('v',4,'x',5,ratio_columns[10:],ws)
    insert_data('y',4,'y',5,ratio_row[10:],ws)

    balance_names = [description[0] for description in conn.execute(f'select * from [{tickers[0]}_balance_sheet]').description][1:]
    balance_rows = curs.execute(f'select * from [{tickers[0]}_balance_sheet]').fetchall()
    insert_data('a',16,'b',51,['']*36, ws)
    insert_data('a',16,'b',16+len(balance_names)-1,balance_names, ws)
    add_table('c',15,[['']*37]*2,ws)
    add_table('c',15,balance_rows,ws)

    income_names = [description[0] for description in conn.execute(f'select * from [{tickers[0]}_income]').description][1:]
    income_rows = curs.execute(f'select * from [{tickers[0]}_income]').fetchall()
    insert_data('h',16,'i',48,['']*33, ws)
    insert_data('h',16,'i',16+len(income_names)-1,income_names, ws)    
    add_table('j',15,[['']*37]*2,ws)
    add_table('j',15,income_rows,ws)

    cash_flow_names = [description[0] for description in conn.execute(f'select * from [{tickers[0]}_cash_flow]').description][1:]
    cash_flow_rows = curs.execute(f'select * from [{tickers[0]}_cash_flow]').fetchall()
    insert_data('o',16,'p',37,['']*22, ws)
    insert_data('o',16,'p',16+len(cash_flow_names)-1,cash_flow_names, ws)
    add_table('q',15,[['']*37]*2,ws)
    add_table('q',15,cash_flow_rows,ws)
    
    compare_table = compare(tickers)

    insert_data('v',15,'w',15+len(compare_table[0])-1,compare_table.pop(0), ws)
    l='x'
    for col in compare_table:
        insert_data(l,15,l,15+len(col)-1,col,ws)
        l=next_ltr(l)
    while l!='ad':
        insert_data(l,15,l,39, ['']*36,ws)
        l=next_ltr(l)

    wb.save(filename)


tickers = list(map(
        lambda query: search_stock_api(query,'s'),
        input('Введите тикеры компаний через запятую: ').replace(' ','').split(',')
        ))
filename =  'book2.xlsx' #input('Введите название файла: ')
update_excel_file(filename, tickers)